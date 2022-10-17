import os
import openpyxl
from openpyxl.styles import Font
import sqlite3
from settings import database_name

filename = 'C:/Users/Tengwei/Desktop/Music/MusicStats.xlsx'
directory = 'C:/Users/Tengwei/Desktop/Music'


# Gets the music score
def get_music_score(song_name: str) -> int:
    with sqlite3.connect(database_name) as con:
        cur = con.cursor()
        cur.execute(
            """
            SELECT score 
              FROM music
             WHERE title = ?""",
            (song_name, )
        )
        score = cur.fetchall()[0][0]
    return score


# Changes the score of a piece of music
def change_music_score(song_name: str, score: int):
    with sqlite3.connect(database_name) as con:
        cur = con.cursor()
        cur.execute(
            """
            UPDATE music
               SET score = ?
             WHERE title = ?""",
            (score, song_name)
        )


# Gets the average score of a genre
def get_average_score(genre: str) -> int:
    with sqlite3.connect(database_name) as con:
        cur = con.cursor()
        cur.execute(
            """
            SELECT AVG(score) 
              FROM music 
             WHERE genre = ?
            """, (genre, )
        )
        average = cur.fetchall()[0][0]
    return average


# # Not sure how this is used currently
# def update_music_list():
#     # Adds the new songs added. If a whole new genre is added it creates a new sheet for the genre
#     global filename
#     global directory
#     wb = openpyxl.load_workbook(filename)
#
#     for sheetname in os.listdir(directory):
#         if os.path.isdir(f'{directory}/{sheetname}'):
#             try:
#                 sheet = wb[sheetname]
#                 average_score = get_average_score(sheetname)
#             except:
#                 sheet = wb.create_sheet(sheetname)
#                 average_score = 20
#             sheet.cell(row=2, column=2).value = "Song Title"
#             sheet.cell(row=2, column=2).font = Font(bold=True)
#             sheet.cell(row=2, column=3).value = "Tracking Score"
#             sheet.cell(row=2, column=3).font = Font(bold=True)
#             for i, file in enumerate(os.listdir(f'C:/Users/Tengwei/Desktop/Music/{sheetname}')):
#                 print(i, file)
#                 music_file_added = False
#                 for i in range(sheet.max_row):
#                     if file == sheet.cell(row=i + 3, column=2).value:
#                         music_file_added = True
#                         break
#
#                 if not music_file_added:
#                     for i in range(sheet.max_row + 1):
#                         if sheet.cell(row=i + 3, column=2).value is None:
#                             sheet.cell(row=i + 3, column=2).value = file
#                             sheet.cell(row=i + 3, column=3).value = average_score
#                             break
#
#     wb.save(filename)
#     wb.close()


# if __name__ == "__main__":
#     update_music_list()
